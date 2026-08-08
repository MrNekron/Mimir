import pandas as pd
import os
import re
import itertools
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill, Font

def _get_safe_filename(name):
    """Очищает название категории для использования в качестве имени файла Windows"""
    return re.sub(r'[\\/*?:"<>|]', "", name).strip()

def _clean_value(value):
    """Принудительно возвращает строку и меняет точки на запятые в числах"""
    if pd.isna(value): return ""
    s = str(value).strip()
    if "." in s:
        parts = s.split('.')
        if len(parts) == 2 and parts[0].replace('-','').isdigit() and parts[1].isdigit():
            return s.replace('.', ',')
    return s

def _get_real_max_column(ws):
    """Находит последнюю колонку, ориентируясь строго на вторую строку (где лежат данные)"""
    max_c = 0
    # Проверяем первые 500 колонок
    for c in range(1, 500):
        if ws.cell(row=2, column=c).value is not None:
            max_c = c
    return max_c

def _get_last_characteristic_column(ws):
    """Находит последний ЧЕТНЫЙ столбец (значение) с данными во второй строке"""
    last_even = 0
    for c in range(2, 500, 2):
        if ws.cell(row=2, column=c).value is not None:
            last_even = c
    return last_even

def save_quasi_to_excel(directory, cat_name, blocks_data, is_auto=False):
    """Основная функция сохранения данных в Excel"""
    try:
        if not cat_name: return False, "Название категории пустое"
        filename = _get_safe_filename(cat_name) + ".xlsx"
        path = os.path.join(directory, filename)
        new_data_rows = []

        # --- 1. ФОРМИРОВАНИЕ ДАННЫХ ---
        if is_auto:
            # РЕЖИМ КОМБИНИРОВАНИЯ (Декартово произведение)
            blocks_with_vals = []
            for idx, (name, vals) in enumerate(blocks_data):
                cleaned = [v for v in [_clean_value(x) for x in vals] if v]
                if cleaned:
                    blocks_with_vals.append({'idx': idx, 'name': name, 'vals': cleaned})
            
            if not blocks_with_vals: return False, "Нет данных для комбинирования"

            combo_gen = itertools.product(*[b['vals'] for b in blocks_with_vals])
            for combination in combo_gen:
                row = [""] * (len(blocks_data) * 2)
                for i, val in enumerate(combination):
                    orig_idx = blocks_with_vals[i]['idx']
                    attr_name = blocks_with_vals[i]['name']
                    row[orig_idx * 2] = attr_name
                    row[orig_idx * 2 + 1] = val
                new_data_rows.append(row)
        else:
            # ОБЫЧНЫЙ РЕЖИМ (Построчно)
            max_rows = max(len(vals) for _, vals in blocks_data) if blocks_data else 0
            for i in range(max_rows):
                row = []
                has_any_val = False
                for attr_name, values_list in blocks_data:
                    val = _clean_value(values_list[i]) if i < len(values_list) else ""
                    if val: has_any_val = True
                    row.extend([attr_name, val])
                if has_any_val:
                    new_data_rows.append(row)

        if not new_data_rows and not os.path.exists(path):
            return False, "Нет данных для создания файла"

        # --- 2. ЗАПИСЬ В EXCEL ---
        if os.path.exists(path):
            try:
                wb = load_workbook(path)
                ws = wb.active
            except:
                wb = Workbook(); ws = wb.active
            start_row = ws.max_row + 1
        else:
            wb = Workbook(); ws = wb.active
            ws.cell(row=1, column=1, value=cat_name)
            start_row = 2

        ws.title = "Лист1"
        plain_font = Font(bold=False, size=10)

        # Записываем данные характеристик
        for r_idx, row_data in enumerate(new_data_rows):
            for c_idx, val in enumerate(row_data):
                cell = ws.cell(row=start_row + r_idx, column=c_idx + 1, value=val)
                cell.font = plain_font

        # Протяжка названий характеристик по всем строкам (включая старые)
        for b_idx, (attr_name, _) in enumerate(blocks_data):
            col_idx = (b_idx * 2) + 1
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=col_idx, value=attr_name)
                cell.font = plain_font

        # --- 3. СТИЛИЗАЦИЯ ШАПКИ ---
        real_width = _get_real_max_column(ws)
        _apply_style(ws, real_width, cat_name)

        try:
            wb.save(path)
        except PermissionError:
            return False, "Для продолжения закройте файл Excel =)"

        return True, (path, len(new_data_rows))

    except Exception as e:
        return False, f"Ошибка Logic: {str(e)}"

def _apply_style(ws, final_data_col, cat_name):
    """Стиль A1: фон #CFE2F3, шрифт 10, объединение минимум до L (12)"""
    fill = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")
    font = Font(color="000000", bold=True, size=10)
    align = Alignment(horizontal="left", vertical="center")
    
    # Записываем название
    ws.cell(row=1, column=1, value=cat_name)

    # Очищаем старые объединения
    if ws.merged_cells:
        for merged_range in list(ws.merged_cells.ranges):
            if merged_range.min_row == 1:
                ws.unmerge_cells(str(merged_range))
    
    # Объединяем ячейку A1: минимум до 12, либо до края данных
    target_merge_end = max(12, final_data_col)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=target_merge_end)
    
    cell = ws["A1"]
    cell.fill = fill
    cell.font = font
    cell.alignment = align

def verify_file_data(path):
    """Функция 'Проверить': сцепка по четным столбцам второй строки с перезаписью"""
    try:
        if not os.path.exists(path):
            return False, "Файл не найден. Сначала сохраните таблицу."

        wb = load_workbook(path)
        ws = wb.active

        # 1. Находим последнюю характеристику (последний ЧЕТНЫЙ столбец)
        last_char_col = _get_last_characteristic_column(ws)
        if last_char_col == 0:
            return False, "Данные характеристик не найдены во 2-й строке."

        # 2. Целевой столбец — нечетный, следующий сразу за последней характеристикой
        target_col = last_char_col + 1

        # 3. Выполняем сцепку для всех строк (со 2-й и ниже)
        for r in range(2, ws.max_row + 1):
            row_values = []
            # Берем только четные столбцы (2, 4, 6...) СЛЕВА от целевого столбца
            for c in range(2, target_col, 2):
                val = ws.cell(row=r, column=c).value
                if val:
                    row_values.append(str(val).strip())
            
            combined_text = " ".join(row_values)
            cleaned_text = " ".join(combined_text.split())
            
            # Записываем (или перезаписываем) результат
            cell = ws.cell(row=r, column=target_col, value=cleaned_text)
            cell.font = Font(bold=False, size=10)

        # 4. Очищаем первую строку в этом столбце (никаких надписей "Проверка")
        ws.cell(row=1, column=target_col, value=None)

        # 5. Обновляем плашку А1 (она расширится до target_col, если он > 12)
        _apply_style(ws, target_col, ws.cell(row=1, column=1).value)

        try:
            wb.save(path)
        except PermissionError:
            return False, "Для продолжения закройте файл Excel =)"

        return True, target_col
    except Exception as e:
        return False, str(e)

def save_quasi_to_csv(directory, category_name, blocks_data, is_auto=False):
    """Логика сохранения в CSV (упрощенная, без стилей)"""
    try:
        filename = _get_safe_filename(category_name) + ".csv"
        path = os.path.join(directory, filename)
        new_rows = []
        
        if is_auto:
            blocks_with_vals = []
            for idx, (name, vals) in enumerate(blocks_data):
                cleaned = [v for v in [_clean_value(x) for x in vals] if v]
                if cleaned: blocks_with_vals.append({'idx': idx, 'name': name, 'vals': cleaned})
            if not blocks_with_vals: return False, "Нет данных"
            combo_gen = itertools.product(*[b['vals'] for b in blocks_with_vals])
            for combination in combo_gen:
                row = [""] * (len(blocks_data) * 2)
                for i, val in enumerate(combination):
                    orig_idx = blocks_with_vals[i]['idx']
                    row[orig_idx * 2], row[orig_idx * 2 + 1] = blocks_with_vals[i]['name'], val
                new_rows.append(row)
        else:
            max_r = max(len(vals) for _, vals in blocks_data) if blocks_data else 0
            for i in range(max_r):
                row = []; has_v = False
                for attr_name, vals in blocks_data:
                    v = _clean_value(vals[i]) if i < len(vals) else ""
                    if v: has_v = True
                    row.extend([attr_name, v])
                if has_v: new_rows.append(row)

        if os.path.exists(path):
            try:
                old_df = pd.read_csv(path, sep=';', encoding='utf-8-sig', header=None, dtype=str).fillna("")
                cat_row, data_df = old_df.iloc[0].tolist(), old_df.iloc[1:]
            except: cat_row, data_df = [category_name], pd.DataFrame()
        else: cat_row, data_df = [category_name], pd.DataFrame()

        final_data_df = pd.concat([data_df, pd.DataFrame(new_rows)], ignore_index=True).fillna("")
        for i, (name, _) in enumerate(blocks_data):
            if i * 2 < final_data_df.shape[1]: final_data_df[i * 2] = name

        with open(path, 'w', encoding='utf-8-sig', newline='') as f:
            width = max(12, final_data_df.shape[1])
            pd.DataFrame([cat_row + [""]*(width - len(cat_row))]).to_csv(f, index=False, header=False, sep=';')
            final_data_df.to_csv(f, index=False, header=False, sep=';')
        return True, (path, len(new_rows))
    except Exception as e:
        return False, str(e)